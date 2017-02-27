import openpyxl
import csv

wb = openpyxl.load_workbook('SpbFishZakaz.xlsx')
lst = []

for Nlist in wb.sheetnames:
    if Nlist != 'Index' and Nlist != 'заказчик' and Nlist != 'скидки' and Nlist != 'Ост' and Nlist != 'прайс' and Nlist != 'скидк' and Nlist != '68':
        ws = wb.get_sheet_by_name(Nlist)
        head = ws['4']


        def SearchRef():
            for ref in head:
                if ref.value == 'Артикул':
                    return ref.column


        def SearchRetailPrice():
            for price in head:
                # if price.value not in ['Рек.розн.', 'Рек.розн. ', 'Рек.розн. Интернет ']:
                if price.value == 'Рек.розн.' or price.value == 'Рек.розн. ' or price.value == 'Рек.розн. Интернет ':
                    return price.column


        def AppToLst():
            for row in range(1, ws.max_row + 1):  # хз почему 1 надо ставить вначале
                ref = ws[SearchRef() + str(row)].value
                price = ws[SearchRetailPrice() + str(row)].value
                if price != None and price != 'Рек.розн.':
                    lst.append([ref, price])


        AppToLst()

with open('Price_NORMARK.csv', 'w', newline='') as csv_r:
    csv_w = csv.writer(csv_r, delimiter=',')
    csv_w.writerows(lst)
